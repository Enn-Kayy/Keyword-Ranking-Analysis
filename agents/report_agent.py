import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime


def generate_excel_report(results):
    """
    Generates a formatted Excel report from the keyword ranking results.
    Highlights rankings visually for easier interpretation.
    """

    # Convert results into a DataFrame with a fixed column order
    df = pd.DataFrame(results)
    df = df[["keyword", "url", "google places", "google links", "page number"]]

    output_path = "data/output/keyword_ranking_report.xlsx"

    # Write data to Excel using openpyxl for formatting support
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Keyword Rankings")

        sheet = writer.book["Keyword Rankings"]

        # Apply conditional formatting to Google Places rank column
        for row in range(2, sheet.max_row + 1):
            rank_cell = sheet[f"C{row}"]

            if rank_cell.value == "Not Found":
                # Highlight missing rankings in red
                rank_cell.fill = PatternFill("solid", fgColor="FF9999")
            elif isinstance(rank_cell.value, int) and rank_cell.value <= 10:
                # Highlight top rankings (Page 1) in green
                rank_cell.fill = PatternFill("solid", fgColor="99FF99")
            else:
                # Highlight lower rankings in yellow
                rank_cell.fill = PatternFill("solid", fgColor="FFFF99")

    print(f"✅ Report generated at {output_path}")
